from bs4 import BeautifulSoup
import datetime
import requests
from openpyxl import Workbook
from openpyxl import load_workbook


## scrape the score
start_ID = 401161209
end_ID = 401161616
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
filename = "score.xlsx"
for i in range(start_ID, end_ID):
    print(i)
    nrows = []
    url = 'https://www.espn.com/nba/game?gameId='+str(i)
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.content, 'html5lib')
    title = soup.title.text
    sheetname1 = title.split(' vs. ')[0]
    mid = title.split(" vs. ")[1].split(" - ")
    sheetname2 = mid[0]
    date = mid[2]
    inner = soup.findAll('span',class_="inner-record")
    if inner:
        hoa1 = inner[0].text[-4:]
        hoa2 = inner[1].text[-4:]
    linescore = soup.find("table",id="linescore")
    if not linescore:
        continue
    tds = linescore.find("tbody").findAll("td")
    quarter = [i.text for i in tds]
    rowHeaders = ['ID','Game_Date','Home_OR_Away','Opponent','First_Quarter_Score_Team','First_Quarter_Score_Opponent','Half_Score_Team','Half_Score_Opponent','Third_Quarter_Team','Third_Quarter_Opponent','Fourth_Quarter_Team','Fourth_Quarter_Opponent','Final_Score_Team','Final_Score_Opponent','ATS']
    rowValues1 = [i, date, hoa1, sheetname2, quarter[1],quarter[7],quarter[2],quarter[8],quarter[3],quarter[9],quarter[4],quarter[10],quarter[5],quarter[11]]
    rowValues2 = [i, date, hoa2, sheetname1, quarter[7],quarter[1],quarter[8],quarter[2],quarter[9],quarter[3],quarter[10],quarter[4],quarter[11],quarter[5]]
    try:
        wb = load_workbook(filename)
        if sheetname1 not in wb.sheetnames:
            ws1 = wb.create_sheet(title=sheetname1)
            ws1.append(rowHeaders)
            ws1.append(rowValues1)
        else:
            ws1 = wb[sheetname1]
            ws1.append(rowValues1)
        if sheetname2 not in wb.sheetnames:
            ws2 = wb.create_sheet(title=sheetname2)
            ws2.append(rowHeaders)
            ws2.append(rowValues2)
        else:
            ws2 = wb[sheetname2]
            ws2.append(rowValues2)
    except FileNotFoundError:
        wb = Workbook()
        ws1 = wb.create_sheet(title=sheetname1)
        ws1.append(rowHeaders)
        ws1.append(rowValues1)
        ws2 = wb.create_sheet(title=sheetname2)
        ws2.append(rowHeaders)
        ws2.append(rowValues2)
    wb.save(filename)
    
    
## scrape the ATS
startdate = datetime.date(2020, 1,10)
wb = load_workbook(filename)
for i in range(62):
    print(i)
    date = startdate + datetime.timedelta(days=i)
    if date.day < 10:
        day = '0' + str(date.day)
    else:
        day = str(date.day)
    url = 'https://www.vegasinsider.com/nba/matchups/matchups.cfm/date/' + '0' + str(date.month) + '-' + day + '-'+'20'#01-09-20'
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.content, 'html5lib')
    divs = soup.findAll('div',class_="SLTables1")
    for div in divs:
        table = div.find('table')
        tds = table.findAll("td", class_="viCellBg2 cellBorderL1 cellTextNorm padLeft")
        tds1 = table.find("td", class_="viHeaderNorm")
        teams = tds1.text.replace("\n","").replace("\t","").split(" @ ")
        ats = [td.text.replace("\n","").replace("\t","") for td in tds if tds.index(td) % 4==3]
        for sheetname in wb.sheetnames:
            if sheetname in teams[0]:
                ws = wb[sheetname]
                dates = [datetime.datetime.strptime(ws['B'][index].value, "%B %d, %Y").date() for index in range(1,len(ws['B']))]
                if date in dates:
                    ws['O' + str(dates.index(date)+2)] = ats[0]
            if sheetname in teams[1]:
                ws = wb[sheetname]
                dates = [datetime.datetime.strptime(ws['B'][index].value, "%B %d, %Y").date() for index in range(1,len(ws['B']))]
                if date in dates:
                    ws['O' + str(dates.index(date)+2)] = ats[1] 
            wb.save('score.xlsx')