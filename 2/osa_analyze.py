from openpyxl import load_workbook
import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from numpy import dot,array,empty_like
from matplotlib.path import Path


def make_path(x1,y1,x2,y2):
    return Path([[x1,y1],[x1,y2],[x2,y2],[x2,y1]])

def perp( a ) :
    b = empty_like(a)
    b[0] = -a[1]
    b[1] = a[0]
    return b


def seg_intersect(a1,a2, b1,b2) :
    da = a2-a1
    db = b2-b1
    dp = a1-b1
    dap = perp(da)
    denom = dot( dap, db)
    num = dot( dap, dp )
   
    x3 = ((num / denom.astype(float))*db + b1)[0]
    y3 = ((num / denom.astype(float))*db + b1)[1]
    p1 = make_path(a1[0],a1[1],a2[0],a2[1])
    p2 = make_path(b1[0],b1[1],b2[0],b2[1])
    if p1.contains_point([x3,y3]) and p2.contains_point([x3,y3]):
        return round(x3,2), round(y3,2)
    else:
        return False

wb = load_workbook('score.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Sheet":
        continue
    print(sheet)
    ws = wb[sheet]
    away = [ws['M'][i].value for i in range(len(ws['M'])) if ws['C'][i].value == "Away"]
    home = [ws['M'][i].value for i in range(len(ws['M'])) if ws['C'][i].value == "Home"]
    away_average = []
    home_average = []
    for i in range(len(away)):
        sum = 0
        for j in range(i+1):
            sum += int(away[j])
        away_average.append(round(sum/(i+1),2))
    for i in range(len(away)):
        sum = 0
        for j in range(i+1):
            sum += int(away[j])
        away_average.append(round(sum/(i+1),2))
    for i in range(len(home)):
        sum = 0
        for j in range(i+1):
            sum += int(home[j])
        home_average.append(round(sum/(i+1),2))
    for i in range(len(home)):
        sum = 0
        for j in range(i+1):
            sum += int(home[j])
        home_average.append(round(sum/(i+1),2))
    marginal = [int(home[i]) for i in range(len(home))]

    optimum_score_x = []
    optimum_score_y = []
    x = range(1,len(home_average)+1)
    for i in range(len(x)-1):
        p1 = array([x[i],home_average[i]])
        p2 = array([x[i+1],home_average[i+1]])
        p3 = array([x[i],marginal[i]])
        p4 = array([x[i+1],marginal[i+1]])
        # print (seg_intersect( p1,p2, p3,p4))
        if seg_intersect( p1,p2, p3,p4):
            optimum_score_x.append(seg_intersect(p1,p2,p3,p4)[0])
            optimum_score_y.append(seg_intersect(p1,p2,p3,p4)[1])

    graph = plt.subplot2grid((8,4), (0,0), colspan=4,rowspan=2)
    graph.set_title("Average home score and Marginal score of" +" "+ sheet,fontsize=20)
    plt.grid(True)
    plt.plot(x,home_average,'g-o', markersize=10, label="Average score")
    plt.plot(range(1,len(marginal)+1),marginal,'r-o',markersize=10,linestyle='dashed',label="Margine Score")
    plt.plot(optimum_score_x,optimum_score_y,'bo',label="Optimum score")
    plt.legend()
    
    table = plt.subplot2grid((8,4), (2,0), colspan=4,rowspan=2)
    s= table.table(cellText=[home_average],colLabels=x,colWidths=[.048]*len(home_average))
    table.axis("off")
    table.set_title("Average home score of Lakers by game numbers", y=0.1)
    s.auto_set_font_size(False)
    s.set_fontsize(12)
    s.scale(1, 1)

    margine_table = plt.subplot2grid((8,4), (4,0), colspan=4,rowspan=2)
    s= margine_table.table(cellText=[marginal],colLabels=x,colWidths=[.048]*len(marginal))
    margine_table.axis("off")
    margine_table.set_title("Margine", y=0.1)
    s.auto_set_font_size(False)
    s.set_fontsize(12)
    s.scale(1,1)

    optimum_table = plt.subplot2grid((8,4), (6,0), colspan=6,rowspan=2)
    s= optimum_table.table(cellText=[optimum_score_y],colLabels=optimum_score_x,colWidths=[.048]*len(optimum_score_x))
    optimum_table.axis("off")
    optimum_table.set_title("Optimum", y=0.3)
    s.auto_set_font_size(False)
    s.set_fontsize(12)
    s.scale(1,1)

    plt.legend()
    fig = plt.figure(1)
    fig.set_size_inches(w=60, h=30)
    fig.subplots_adjust(hspace=0.9)

    plt.show()

